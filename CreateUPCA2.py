import tkinter as tk
from tkinter import filedialog
import os
import xlrd
from PIL import Image, ImageDraw, ImageFont
import barcode
from barcode.writer import ImageWriter
from openpyxl import load_workbook

# Function to generate barcode image with Item Number and Item Description
def generate_barcode_with_text(data, item_number, item_description, filename):
    code128 = barcode.get_barcode_class('upca')
    barcode_instance = code128(data, writer=ImageWriter())
    barcode_image = barcode_instance.render()

    # Load font for text
    font_path = "C:\\ClairProjetKemp\\KempUPC\\fontcodeupceh3_tr.ttf"
    font_size = 18
    try:
        font = ImageFont.truetype(font_path, font_size)
    except IOError:
        print(f"Font file not found: {font_path}")
        exit()

    # Create a new image with enough space for the barcode, item number, and item description
    barcode_width, barcode_height = barcode_image.size
    total_height = barcode_height + 2 * font_size + 10  # 10 pixels padding between barcode and text
    composite_image = Image.new('RGB', (barcode_width, total_height), color='white')

    # Paste barcode onto the composite image
    composite_image.paste(barcode_image, (0, 0))

    # Draw Item Number below the barcode
    draw = ImageDraw.Draw(composite_image)
    item_number_width = font_size * len(item_number) // 2  # Rough estimate
    draw.text(((barcode_width - item_number_width) / 2, barcode_height + 5), item_number, fill='black', font=font)

    # Draw Item Description centered above the barcode
    item_description_width = font_size * len(item_description) // 2  # Rough estimate
    draw.text(((barcode_width - item_description_width) / 2, barcode_height - font_size - 5),
              item_description, fill='black', font=font)
    # Save composite image
    save_path = os.path.join("C:\\Temp", filename)
    composite_image.save(save_path)

# Function to browse for Excel file
def browse_excel_file():
    root = tk.Tk()
    root.withdraw()  # Hide the main window
    file_path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx;*.xls")])
    return file_path

# Get the path of the Excel file
excel_file_path = browse_excel_file()

# Check if a file was selected
if not excel_file_path:
    print("No file selected. Exiting...")
    exit()

# Load Excel spreadsheet
_, file_extension = os.path.splitext(excel_file_path)
if file_extension == '.xlsx':
    wb = load_workbook(filename=excel_file_path, read_only=True)
elif file_extension == '.xls':
    print("Unsupported file format. Exiting...")
    exit()

sheet = wb.active  # Assuming data is in the first sheet

# Generate barcode for each row in the spreadsheet
for row in sheet.iter_rows(min_row=2):  # Assuming the first row contains headers
    item_number = str(row[0].value)  # Assuming Item Numbers are in the first column
    item_description = str(row[1].value)  # Assuming Item Descriptions are in the second column
    upc_code = '0' + str(row[3].value)  # Assuming UPC codes are in the fourth column, add leading zero
    barcode_filename = f"{upc_code}.png"
    generate_barcode_with_text(upc_code, item_number, item_description, barcode_filename)
    print(f"Barcode generated for {upc_code}")

print("All barcodes generated successfully.")
